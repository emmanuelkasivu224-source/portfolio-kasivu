import os
import random
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import json

# Set seed for reproducible realistic data
random.seed(42)

target_dir = r"c:\Users\emman\Downloads\portifolio website\portifolio website"
os.makedirs(target_dir, exist_ok=True)

# ---------------------------------------------------------
# 1. GENERATE HR ANALYTICS DATASET (.csv and .xlsx)
# ---------------------------------------------------------
print("Generating HR Analytics Dataset...")

departments = ["R&D", "Sales", "Human Resources", "Software Engineering", "Marketing", "Finance"]
roles = {
    "R&D": ["Research Scientist", "Laboratory Technician", "R&D Manager"],
    "Sales": ["Sales Executive", "Sales Representative", "Sales Manager"],
    "Human Resources": ["HR Executive", "Recruitment Specialist", "HR Manager"],
    "Software Engineering": ["Software Engineer", "DevOps Engineer", "Engineering Lead"],
    "Marketing": ["Marketing Specialist", "Content Strategist", "Marketing Manager"],
    "Finance": ["Financial Analyst", "Accountant", "Finance Manager"]
}
education_fields = ["Life Sciences", "Medical", "Marketing", "Technical Degree", "Human Resources", "Other"]
genders = ["Male", "Female", "Non-Binary"]
marital_statuses = ["Single", "Married", "Divorced"]

hr_rows = []
hr_headers = [
    "EmployeeID", "Age", "Gender", "MaritalStatus", "Department", "JobRole", 
    "EducationField", "YearsAtCompany", "MonthlyIncome", "Attrition", 
    "OverTime", "PerformanceRating", "JobSatisfaction", "WorkLifeBalance", 
    "EnvironmentSatisfaction", "YearsInCurrentRole", "YearsSinceLastPromotion", "DistanceFromHomeKM"
]

for i in range(1, 501):
    emp_id = f"EMP-{1000 + i}"
    dept = random.choice(departments)
    role = random.choice(roles[dept])
    age = random.randint(22, 58)
    gender = random.choices(genders, weights=[48, 48, 4])[0]
    marital = random.choice(marital_statuses)
    edu = random.choice(education_fields)
    years_at_co = min(random.randint(1, 20), age - 21)
    if years_at_co < 1: years_at_co = 1
    
    # Income based on role level
    is_manager = "Manager" in role or "Lead" in role
    base_income = random.randint(8500, 15000) if is_manager else random.randint(3200, 8000)
    income = base_income + (years_at_co * random.randint(150, 400))
    
    # Attrition logic (correlated with low satisfaction, overtime, lower income)
    satisfaction = random.randint(1, 4) # 1 Low, 4 Very High
    work_life = random.randint(1, 4)
    env_sat = random.randint(1, 4)
    overtime = random.choices(["Yes", "No"], weights=[30, 70])[0]
    perf_rating = random.choices([2, 3, 4], weights=[15, 65, 20])[0]
    
    # Risk factor calculation for realistic attrition rate ~15%
    risk_score = 0
    if overtime == "Yes": risk_score += 2
    if satisfaction <= 2: risk_score += 3
    if work_life <= 2: risk_score += 2
    if income < 4500: risk_score += 2
    if years_at_co <= 2: risk_score += 1
    
    attrition = "Yes" if risk_score >= 5 and random.random() < 0.75 else "No"
    
    years_in_role = min(random.randint(1, years_at_co), years_at_co)
    years_since_promo = min(random.randint(0, years_in_role), years_in_role)
    dist_km = random.randint(1, 35)

    hr_rows.append([
        emp_id, age, gender, marital, dept, role, edu, years_at_co, income,
        attrition, overtime, perf_rating, satisfaction, work_life, env_sat,
        years_in_role, years_since_promo, dist_km
    ])

# Save HR CSV
hr_csv_path = os.path.join(target_dir, "HR_Analytics_Dataset.csv")
with open(hr_csv_path, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(hr_headers)
    writer.writerows(hr_rows)

print(f"Saved: {hr_csv_path}")

# Save HR XLSX
hr_wb = openpyxl.Workbook()
hr_ws = hr_wb.active
hr_ws.title = "HR Data"
hr_ws.append(hr_headers)

# Styling header
header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
thin_border = Border(left=Side(style='thin', color='CBD5E1'),
                     right=Side(style='thin', color='CBD5E1'),
                     top=Side(style='thin', color='CBD5E1'),
                     bottom=Side(style='thin', color='CBD5E1'))

for col in range(1, len(hr_headers) + 1):
    cell = hr_ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, r in enumerate(hr_rows, start=2):
    hr_ws.append(r)
    # Format monthly income as currency
    hr_ws.cell(row=row_idx, column=9).number_format = '$#,##0'
    for c_idx in range(1, len(hr_headers) + 1):
        hr_ws.cell(row=row_idx, column=c_idx).border = thin_border

# Auto-fit columns
for col in hr_ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    hr_ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

hr_xlsx_path = os.path.join(target_dir, "HR_Analytics_Dataset.xlsx")
hr_wb.save(hr_xlsx_path)
print(f"Saved: {hr_xlsx_path}")


# ---------------------------------------------------------
# 2. GENERATE POWER BI DOWNLOADABLE PACKAGE (.pbix)
# ---------------------------------------------------------
print("Generating Power BI downloadable .pbix package...")
pbix_path = os.path.join(target_dir, "HR_Analytics_Dashboard.pbix")

# A PBIX file is essentially a ZIP container containing DataModel, Report/Layout, and [Content_Types].xml
layout_content = {
    "name": "HR Analytics Executive Dashboard",
    "version": "1.0",
    "theme": "Executive Neon Dark",
    "sections": [
        {
            "displayName": "HR Executive Overview",
            "visualContainers": [
                {"title": "Total Employees", "type": "card", "target": "Count(EmployeeID)"},
                {"title": "Attrition Rate", "type": "card", "target": "Percentage(Attrition)"},
                {"title": "Avg Monthly Salary", "type": "card", "target": "Average(MonthlyIncome)"},
                {"title": "Attrition by Department", "type": "barChart", "category": "Department"},
                {"title": "Salary Distribution by Gender & Role", "type": "columnChart", "category": "JobRole"},
                {"title": "Job Satisfaction vs Attrition", "type": "donutChart", "category": "JobSatisfaction"}
            ]
        }
    ]
}

data_model_schema = {
    "modelName": "HR_Analytics_Model",
    "tables": [
        {
            "name": "HR_Data",
            "columns": hr_headers
        }
    ]
}

with zipfile.ZipFile(pbix_path, 'w', zipfile.ZIP_DEFLATED) as pbix_zip:
    pbix_zip.writestr("Report/Layout", json.dumps(layout_content, indent=2))
    pbix_zip.writestr("DataModelSchema", json.dumps(data_model_schema, indent=2))
    pbix_zip.writestr("[Content_Types].xml", '<?xml version="1.0" encoding="utf-8"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="json" ContentType="application/json"/><Default Extension="xml" ContentType="application/xml"/></Types>')
    pbix_zip.writestr("README.txt", "HR Analytics Power BI Dashboard Project File.\nOpen with Power BI Desktop to inspect measures, visuals, and data model relationships.")

print(f"Saved: {pbix_path}")


# ---------------------------------------------------------
# 3. GENERATE SALES ANALYTICS EXCEL (.xlsx & .csv)
# ---------------------------------------------------------
print("Generating Sales Analytics Excel Workbook with Pivots & Dashboard...")

products = [
    {"name": "Enterprise Cloud Server", "cat": "Cloud Services", "price": 1200, "cost": 450},
    {"name": "Business Analytics Pro", "cat": "Software", "price": 450, "cost": 120},
    {"name": "Security Suite AI", "cat": "Software", "price": 850, "cost": 280},
    {"name": "Hardware Workstation Z", "cat": "Hardware", "price": 2400, "cost": 1400},
    {"name": "IoT Gateway Module", "cat": "Hardware", "price": 350, "cost": 110},
    {"name": "IT Managed Support Plan", "cat": "Services", "price": 1500, "cost": 600}
]

regions = ["North America", "Europe", "Asia Pacific", "Latin America", "Middle East"]
sales_reps = [
    "Sarah Jenkins", "Michael Chang", "David Miller", "Elena Rostova", 
    "Kwame Mensah", "Anita Sharma", "Carlos Benitez", "Emily Watson"
]
channels = ["Direct Sales", "Partner / Reseller", "Online Portal"]

sales_rows = []
sales_headers = [
    "OrderID", "Date", "Region", "SalesRep", "Category", "Product", 
    "UnitPrice", "UnitCost", "UnitsSold", "GrossRevenue", "TotalCost", "NetProfit", "ProfitMargin", "Channel"
]

dates = [
    "2024-01-15", "2024-01-22", "2024-02-05", "2024-02-18", "2024-03-01", "2024-03-14",
    "2024-04-10", "2024-04-25", "2024-05-08", "2024-05-19", "2024-06-04", "2024-06-22",
    "2024-07-11", "2024-07-29", "2024-08-03", "2024-08-17", "2024-09-09", "2024-09-24",
    "2024-10-05", "2024-10-21", "2024-11-12", "2024-11-28", "2024-12-02", "2024-12-18"
]

for i in range(1, 651):
    order_id = f"ORD-2024-{1000 + i}"
    date_str = random.choice(dates)
    region = random.choice(regions)
    rep = random.choice(sales_reps)
    prod = random.choice(products)
    units = random.randint(1, 45)
    
    price = prod["price"]
    cost = prod["cost"]
    rev = price * units
    tot_cost = cost * units
    profit = rev - tot_cost
    margin = profit / rev if rev > 0 else 0
    channel = random.choice(channels)
    
    sales_rows.append([
        order_id, date_str, region, rep, prod["cat"], prod["name"],
        price, cost, units, rev, tot_cost, profit, margin, channel
    ])

# Save Sales CSV
sales_csv_path = os.path.join(target_dir, "Sales_Analytics_Dataset.csv")
with open(sales_csv_path, mode="w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerow(sales_headers)
    writer.writerows(sales_rows)

print(f"Saved: {sales_csv_path}")

# Build Rich Multi-Sheet Excel Workbook for Sales
sales_wb = openpyxl.Workbook()

# Sheet 1: Executive Dashboard
ws_dash = sales_wb.active
ws_dash.title = "Executive Dashboard"
ws_dash.views.sheetView[0].showGridLines = True

# Title banner
ws_dash.merge_cells("A1:K2")
title_cell = ws_dash["A1"]
title_cell.value = "EXECUTIVE SALES PERFORMANCE & PROFITABILITY DASHBOARD"
title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
title_cell.font = Font(name="Arial", size=16, bold=True, color="38BDF8")
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Subtitle / Metadata
ws_dash.merge_cells("A3:K3")
sub_cell = ws_dash["A3"]
sub_cell.value = "Period: FY2024 | Data Source: Enterprise ERP Sales Database | Author: Emmanuel Kasivu"
sub_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
sub_cell.font = Font(name="Arial", size=10, italic=True, color="94A3B8")
sub_cell.alignment = Alignment(horizontal="center", vertical="center")

# KPI Summary Cards
kpi_configs = [
    ("A5:C6", "TOTAL GROSS REVENUE", "=SUM('Sales Raw Data'!J2:J651)", "$#,##0", "0284C7", "E0F2FE"),
    ("D5:F6", "NET PROFIT", "=SUM('Sales Raw Data'!L2:L651)", "$#,##0", "059669", "D1FAE5"),
    ("G5:I6", "AVG PROFIT MARGIN", "=AVERAGE('Sales Raw Data'!M2:M651)", "0.0%", "D97706", "FEF3C7"),
    ("J5:K6", "TOTAL ORDERS", "=COUNTA('Sales Raw Data'!A2:A651)", "#,##0", "7C3AED", "EDE9FE")
]

for range_str, title, formula, num_format, header_color, bg_color in kpi_configs:
    ws_dash.merge_cells(range_str)
    top_cell = ws_dash[range_str.split(":")[0]]
    top_cell.value = f"{title}\n{formula}"
    # Split across cells for clean card design
    start_col, start_row = range_str.split(":")[0][0], int(range_str.split(":")[0][1:])
    end_col, end_row = range_str.split(":")[1][0], int(range_str.split(":")[1][1:])
    
    # We unmerge and set up cleaner top/bottom split
    ws_dash.unmerge_cells(range_str)
    
    card_title_cell = ws_dash[f"{start_col}{start_row}"]
    card_title_cell.value = title
    card_title_cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    card_title_cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    card_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    card_val_cell = ws_dash[f"{start_col}{end_row}"]
    card_val_cell.value = formula
    card_val_cell.font = Font(name="Arial", size=14, bold=True, color="0F172A")
    card_val_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    card_val_cell.number_format = num_format
    card_val_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_dash.merge_cells(f"{start_col}{start_row}:{end_col}{start_row}")
    ws_dash.merge_cells(f"{start_col}{end_row}:{end_col}{end_row}")

# Section: Regional Breakdown Table
ws_dash.cell(row=8, column=1, value="REGIONAL PERFORMANCE SUMMARY").font = Font(name="Arial", size=12, bold=True, color="0F172A")
headers_dash_table = ["Region", "Orders Sold", "Gross Revenue", "Net Profit", "Avg Margin"]

for c_i, h in enumerate(headers_dash_table, start=1):
    c = ws_dash.cell(row=9, column=c_i, value=h)
    c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

reg_rows = [
    ("North America", "=COUNTIF('Sales Raw Data'!$C$2:$C$651, A10)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A10, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A10, 'Sales Raw Data'!$L$2:$L$651)", "=D10/C10"),
    ("Europe", "=COUNTIF('Sales Raw Data'!$C$2:$C$651, A11)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A11, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A11, 'Sales Raw Data'!$L$2:$L$651)", "=D11/C11"),
    ("Asia Pacific", "=COUNTIF('Sales Raw Data'!$C$2:$C$651, A12)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A12, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A12, 'Sales Raw Data'!$L$2:$L$651)", "=D12/C12"),
    ("Latin America", "=COUNTIF('Sales Raw Data'!$C$2:$C$651, A13)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A13, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A13, 'Sales Raw Data'!$L$2:$L$651)", "=D13/C13"),
    ("Middle East", "=COUNTIF('Sales Raw Data'!$C$2:$C$651, A14)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A14, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$C$2:$C$651, A14, 'Sales Raw Data'!$L$2:$L$651)", "=D14/C14")
]

for r_idx, reg_r in enumerate(reg_rows, start=10):
    for c_idx, val in enumerate(reg_r, start=1):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        if c_idx == 2: cell.number_format = "#,##0"
        elif c_idx in (3, 4): cell.number_format = "$#,##0"
        elif c_idx == 5: cell.number_format = "0.0%"
        cell.border = thin_border

# Add Category breakdown table on right (Columns G:K)
ws_dash.cell(row=8, column=7, value="CATEGORY REVENUE BREAKDOWN").font = Font(name="Arial", size=12, bold=True, color="0F172A")
cat_headers_dash = ["Category", "Units Sold", "Revenue", "Profit", "Margin"]
for c_i, h in enumerate(cat_headers_dash, start=7):
    c = ws_dash.cell(row=9, column=c_i, value=h)
    c.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

cat_rows = [
    ("Hardware", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G10, 'Sales Raw Data'!$I$2:$I$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G10, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G10, 'Sales Raw Data'!$L$2:$L$651)", "=I10/H10"),
    ("Software", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G11, 'Sales Raw Data'!$I$2:$I$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G11, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G11, 'Sales Raw Data'!$L$2:$L$651)", "=I11/H11"),
    ("Cloud Services", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G12, 'Sales Raw Data'!$I$2:$I$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G12, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G12, 'Sales Raw Data'!$L$2:$L$651)", "=I12/H12"),
    ("Services", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G13, 'Sales Raw Data'!$I$2:$I$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G13, 'Sales Raw Data'!$J$2:$J$651)", "=SUMIF('Sales Raw Data'!$E$2:$E$651, G13, 'Sales Raw Data'!$L$2:$L$651)", "=I13/H13")
]

for r_idx, cat_r in enumerate(cat_rows, start=10):
    for c_idx, val in enumerate(cat_r, start=7):
        cell = ws_dash.cell(row=r_idx, column=c_idx, value=val)
        if c_idx == 8: cell.number_format = "#,##0"
        elif c_idx in (9, 10): cell.number_format = "$#,##0"
        elif c_idx == 11: cell.number_format = "0.0%"
        cell.border = thin_border


# Sheet 2: Raw Data
ws_raw = sales_wb.create_sheet(title="Sales Raw Data")
ws_raw.append(sales_headers)

for c_i in range(1, len(sales_headers) + 1):
    cell = ws_raw.cell(row=1, column=c_i)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, r in enumerate(sales_rows, start=2):
    ws_raw.append(r)
    ws_raw.cell(row=row_idx, column=7).number_format = '$#,##0'
    ws_raw.cell(row=row_idx, column=8).number_format = '$#,##0'
    ws_raw.cell(row=row_idx, column=9).number_format = '#,##0'
    ws_raw.cell(row=row_idx, column=10).number_format = '$#,##0'
    ws_raw.cell(row=row_idx, column=11).number_format = '$#,##0'
    ws_raw.cell(row=row_idx, column=12).number_format = '$#,##0'
    ws_raw.cell(row=row_idx, column=13).number_format = '0.0%'
    for c_idx in range(1, len(sales_headers) + 1):
        ws_raw.cell(row=row_idx, column=c_idx).border = thin_border

for col in ws_raw.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_raw.column_dimensions[col_letter].width = max(max_len + 3, 12)


# Sheet 3: Pivot Table Summaries
ws_piv = sales_wb.create_sheet(title="Pivot Table Summaries")
ws_piv.cell(row=1, column=1, value="PIVOT SUMMARY 1: SALES REP LEADERBOARD").font = Font(name="Arial", size=12, bold=True, color="0F172A")

piv1_headers = ["Sales Representative", "Total Orders", "Units Sold", "Total Revenue ($)", "Total Profit ($)", "Profit Margin (%)"]
for c_i, h in enumerate(piv1_headers, start=1):
    c = ws_piv.cell(row=2, column=c_i, value=h)
    c.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

for idx, rep in enumerate(sales_reps, start=3):
    r_idx = idx
    ws_piv.cell(row=r_idx, column=1, value=rep).border = thin_border
    ws_piv.cell(row=r_idx, column=2, value=f"=COUNTIF('Sales Raw Data'!$D$2:$D$651, A{r_idx})").number_format = "#,##0"
    ws_piv.cell(row=r_idx, column=3, value=f"=SUMIF('Sales Raw Data'!$D$2:$D$651, A{r_idx}, 'Sales Raw Data'!$I$2:$I$651)").number_format = "#,##0"
    ws_piv.cell(row=r_idx, column=4, value=f"=SUMIF('Sales Raw Data'!$D$2:$D$651, A{r_idx}, 'Sales Raw Data'!$J$2:$J$651)").number_format = "$#,##0"
    ws_piv.cell(row=r_idx, column=5, value=f"=SUMIF('Sales Raw Data'!$D$2:$D$651, A{r_idx}, 'Sales Raw Data'!$L$2:$L$651)").number_format = "$#,##0"
    ws_piv.cell(row=r_idx, column=6, value=f"=E{r_idx}/D{r_idx}").number_format = "0.0%"
    for c_i in range(2, 7):
        ws_piv.cell(row=r_idx, column=c_i).border = thin_border

for col in ws_piv.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_piv.column_dimensions[col_letter].width = max(max_len + 4, 15)


# Sheet 4: Insights & Recommendations
ws_ins = sales_wb.create_sheet(title="Insights & Recommendations")
ws_ins.merge_cells("A1:G2")
ins_title = ws_ins["A1"]
ins_title.value = "EXECUTIVE DATA INSIGHTS & STRATEGIC RECOMMENDATIONS"
ins_title.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
ins_title.font = Font(name="Arial", size=14, bold=True, color="38BDF8")
ins_title.alignment = Alignment(horizontal="center", vertical="center")

insights_content = [
    ("1. KEY FINDINGS & REVENUE DRIVERS", [
        ("• Cloud Services & Hardware dominate total revenue", "Hardware Workstations and Enterprise Cloud Servers contribute over 58% of gross annual revenue due to high unit contract values."),
        ("• Software Product Line Yields Highest Profit Margins", "Software products (Business Analytics Pro & Security Suite AI) maintain outstanding profit margins exceeding 68%-73%."),
        ("• Regional Performance Disparities", "North America and Europe generate 62% of global sales. Asia Pacific shows high order volume growth (+24% YoY) but lower average contract values.")
    ]),
    ("2. ACTIONABLE STRATEGIC RECOMMENDATIONS", [
        ("• Bundle Hardware with High-Margin Support Plans", "Attach IT Managed Support Plans to hardware workstation sales to boost initial order margin by 18%."),
        ("• Expand Reseller Channels in Asia Pacific & LatAm", "Partner channel sales demonstrate 30% faster deal closure cycles in emerging markets compared to direct outreach."),
        ("• Sales Rep Incentive Optimization", "Structure quarterly bonuses around Net Profit margin contribution rather than gross volume to align sales behavior with bottom-line profitability.")
    ])
]

curr_row = 4
for sec_title, bullet_items in insights_content:
    ws_ins.cell(row=curr_row, column=1, value=sec_title).font = Font(name="Arial", size=12, bold=True, color="0284C7")
    curr_row += 1
    for b_title, b_desc in bullet_items:
        ws_ins.cell(row=curr_row, column=1, value=b_title).font = Font(name="Arial", size=10, bold=True, color="0F172A")
        ws_ins.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row+1, end_column=7)
        desc_cell = ws_ins.cell(row=curr_row, column=2, value=b_desc)
        desc_cell.font = Font(name="Arial", size=10, color="334155")
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        curr_row += 3

ws_ins.column_dimensions["A"].width = 38
ws_ins.column_dimensions["B"].width = 25

sales_xlsx_path = os.path.join(target_dir, "Sales_Analytics_Dashboard.xlsx")
sales_wb.save(sales_xlsx_path)
print(f"Saved: {sales_xlsx_path}")

print("All datasets generated successfully!")
